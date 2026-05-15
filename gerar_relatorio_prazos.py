"""
Gerador de Relatório de Prazos Preclusivos — Versão Web (PAINEL_WEB)
Documentação: DOCUMENTAÇÃO COMPLETA — RELATÓRIO DE PRAZOS PRECLUSIVOS
"""

import io, re, sys, unicodedata
from datetime import date, datetime, timedelta
from openpyxl.styles.borders import Border, Side

import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

try:
    import pandas as pd
    _USE_PANDAS = True
except ImportError:
    _USE_PANDAS = False

try:
    import holidays as holidays_lib
    _USE_HOLIDAYS = True
except ImportError:
    _USE_HOLIDAYS = False

# ─── CORES (ARGB conforme documentação) ──────────────────────────────────────
F_VERDE    = PatternFill('solid', fgColor='C6EFCE')
F_AMARELO  = PatternFill('solid', fgColor='FFEB9C')
F_ROSA     = PatternFill('solid', fgColor='FFC7CE')
F_VERMELHO = PatternFill('solid', fgColor='CC0000')
F_LARANJA  = PatternFill('solid', fgColor='FFCC99')
F_CABEC    = PatternFill('solid', fgColor='2F4F8F')
F_TOTAL    = PatternFill('solid', fgColor='D9E1F2')
F_DATA_ERR = PatternFill('solid', fgColor='FF0000')
F_NONE     = PatternFill(fill_type=None)

FONTE_BRANCA  = Font(color='FFFFFF', bold=True, size=11)
FONTE_CABEC   = Font(color='FFFFFF', bold=True, size=11)
FONTE_NEGRITO = Font(bold=True, size=11)
FONTE_TITULO  = Font(bold=True, size=13)
FONTE_NORMAL  = Font(size=11)

AL_C       = Alignment(horizontal='center', vertical='center', wrap_text=False)
AL_E       = Alignment(horizontal='left',   vertical='center', wrap_text=False)
AL_C_WRAP  = Alignment(horizontal='center', vertical='center', wrap_text=True)
AL_E_WRAP  = Alignment(horizontal='left',   vertical='center', wrap_text=True)

BORDA_FINA = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin')
)

# ─── FERIADOS NACIONAIS (sem Carnaval e Corpus Christi) ──────────────────────
def _feriados_np(anos):
    if not _USE_HOLIDAYS:
        return np.array([], dtype='datetime64[D]')
    br = holidays_lib.Brazil(years=anos)
    excluir = [d for d, n in list(br.items())
               if 'carnaval' in n.lower() or 'corpus' in n.lower()]
    for d in excluir:
        del br[d]
    return np.array([d.strftime('%Y-%m-%d') for d in sorted(br.keys())],
                    dtype='datetime64[D]')

# ─── MAPEAMENTO EXECUTANTE → RESPONSÁVEL ─────────────────────────────────────
# Nomes completos conforme documentação. Comparação normalizada (sem acentos/case).
RESPONSAVEL_MAP = {
    # HELANZIA
    "Wellington Pereira da Rocha Filho":                   "HELANZIA DE ARAUJO XAVIER WICHMANN",
    "Keliane de Oliveira":                                 "HELANZIA DE ARAUJO XAVIER WICHMANN",
    "Natalia Paiva de Paula":                              "HELANZIA DE ARAUJO XAVIER WICHMANN",
    "Gustavo Lopes Alencar Filho":                         "HELANZIA DE ARAUJO XAVIER WICHMANN",
    "Artur Saraiva de Andrade":                            "HELANZIA DE ARAUJO XAVIER WICHMANN",
    "Victor Emanoel Fradique Accioly Fontenele":           "HELANZIA DE ARAUJO XAVIER WICHMANN",
    "Helanzia de Araujo Xavier Wichmann":                  "HELANZIA DE ARAUJO XAVIER WICHMANN",
    "Helanzia de Araujo Xavier Wichamnn":                  "HELANZIA DE ARAUJO XAVIER WICHMANN",
    "Roberta Rayanne Vasconcelos Boto":                    "HELANZIA DE ARAUJO XAVIER WICHMANN",
    # LUCIANE
    "Matheus Cavalcanti de Araujo":                        "LUCIANE MODERNEL MENDES",
    "Layla Evelyn Nascimento Pinheiro":                    "LUCIANE MODERNEL MENDES",
    "Antonio Eduardo Goes Aguiar Filho":                   "LUCIANE MODERNEL MENDES",
    "Eduardo Blasques Martine":                            "LUCIANE MODERNEL MENDES",
    "Sane Borges Borgomoni":                               "LUCIANE MODERNEL MENDES",
    "Luciane Modernel Mendes":                             "LUCIANE MODERNEL MENDES",
    "Erika Paula Santos Lima":                             "LUCIANE MODERNEL MENDES",
    "Erika Paula":                                         "LUCIANE MODERNEL MENDES",
    # GABRIEL
    "Juliana de Oliveira Rocha":                           "GABRIEL GIORGIO CICCHELERO",
    "Alysson Narbal de Oliveira Sombra":                   "GABRIEL GIORGIO CICCHELERO",
    "Jamile Barreto":                                      "GABRIEL GIORGIO CICCHELERO",
    "Gabriel Giorgio Cicchelero":                          "GABRIEL GIORGIO CICCHELERO",
    "Irene Flavia Serenario":                              "GABRIEL GIORGIO CICCHELERO",
    "Irene Flávia Serenário":                              "GABRIEL GIORGIO CICCHELERO",
    "Armando Helio Almeida Monteiro de Moraes":            "GABRIEL GIORGIO CICCHELERO",
    "Armando Hélio Almeida Monteiro de Moraes":            "GABRIEL GIORGIO CICCHELERO",
    # JENIFFER
    "Paulo Marcio Soares de Carvalho Filho":               "JENIFFER ROSA BARBOSA DE SALES",
    "Jeniffer Rosa Barbosa de Sales":                      "JENIFFER ROSA BARBOSA DE SALES",
    # JULIANA MIRELLA
    "Thallys Anderson Ferreira de Lima":                   "JULIANA MIRELLA ALVES RODRIGUES",
    "Juliana Mirella Alves Rodrigues":                     "JULIANA MIRELLA ALVES RODRIGUES",
    # NAYANDERSON
    "Andre Viana Garrido":                                 "NAYANDERSON LUAN MELLO PINHEIRO",
    "Emerson Travassos Torquato":                          "NAYANDERSON LUAN MELLO PINHEIRO",
    "Nayanderson Luan Mello Pinheiro":                     "NAYANDERSON LUAN MELLO PINHEIRO",
    "Yuri Gondim de Amorim":                               "NAYANDERSON LUAN MELLO PINHEIRO",
    # YURI
    "Luiz Guilherme Goncalves Girao":                      "YURI ALVES BARROS DOS SANTOS",
    "Luiz Guilherme Gonçalves Girão":                      "YURI ALVES BARROS DOS SANTOS",
    "Yuri Alves Barros dos Santos":                        "YURI ALVES BARROS DOS SANTOS",
    # MARCELLE
    "Mariana Mota Frota":                                  "MARCELLE LEITE RENTROIA",
    "Marcelle Leite Rentroia":                             "MARCELLE LEITE RENTROIA",
    # SUZANA
    "Giovanna Campos Pereira":                             "SUZANA MARIA CAMPOS MARANHAO DE LIMA",
    "Giovanna Cesar Ferreira":                             "SUZANA MARIA CAMPOS MARANHAO DE LIMA",
    "Daniel Barros de Oliveira":                           "SUZANA MARIA CAMPOS MARANHAO DE LIMA",
    "Leticia Oliveira da Silva":                           "SUZANA MARIA CAMPOS MARANHAO DE LIMA",
    "Tatiane Carmo Santa Rosa":                            "SUZANA MARIA CAMPOS MARANHAO DE LIMA",
    "Francoise Catherine Souza Alves":                     "SUZANA MARIA CAMPOS MARANHAO DE LIMA",
    "Evilany Gabriela Braga Pontes":                       "SUZANA MARIA CAMPOS MARANHAO DE LIMA",
    "Suzana Maria Campos Maranhao de Lima":                "SUZANA MARIA CAMPOS MARANHAO DE LIMA",
    "Suzana Maria Campos Maranhão de Lima":                "SUZANA MARIA CAMPOS MARANHAO DE LIMA",
    # TICIANNA
    "Ticianna Pires de Souza":                             "TICIANNA PIRES DE SOUZA",
    # RONALD
    "Alexia Alencar Capibaribe":                           "RONALD FEITOSA AGUIAR FILHO",
    "Ronald Feitosa Aguiar Filho":                         "RONALD FEITOSA AGUIAR FILHO",
}

ORDEM_SUPERVISORES = [
    "GABRIEL GIORGIO CICCHELERO",
    "HELANZIA DE ARAUJO XAVIER WICHMANN",
    "JENIFFER ROSA BARBOSA DE SALES",
    "JULIANA MIRELLA ALVES RODRIGUES",
    "LUCIANE MODERNEL MENDES",
    "MARCELLE LEITE RENTROIA",
    "NAYANDERSON LUAN MELLO PINHEIRO",
    "RONALD FEITOSA AGUIAR FILHO",
    "SUZANA MARIA CAMPOS MARANHAO DE LIMA",
    "TICIANNA PIRES DE SOUZA",
    "YURI ALVES BARROS DOS SANTOS",
]

# Colunas de saída nas abas de supervisor (ordem e larguras conforme documentação)
COLUNAS_SAIDA = [
    ('Data',              18),
    ('Executante',        35),
    ('Responsavel',       38),
    ('Natureza',          18),
    ('Tipo / Subtipo',    25),
    ('Descricao',         40),
    ('Processo',          28),
    ('Cliente',           35),
    ('Contrario',         35),
    ('Orgao / Cidade',    20),
    ('Vara',              15),
    ('Status',            15),
    ('Justificativa',     50),
    ('Pasta',             15),
    ('Divergencia de Data', 60),
]
NOMES_COLUNAS = [c[0] for c in COLUNAS_SAIDA]
LARGURAS_COLUNAS = [c[1] for c in COLUNAS_SAIDA]

# ─── HELPERS ─────────────────────────────────────────────────────────────────
def _norm(s):
    nfd = unicodedata.normalize('NFD', str(s or '').strip().upper())
    return ''.join(c for c in nfd if unicodedata.category(c) != 'Mn')

def _mapear(nome_raw):
    if not nome_raw or str(nome_raw).strip() in ('', 'nan', 'None'):
        return None
    nome = str(nome_raw).split('\n')[0].strip()
    # exact
    if nome in RESPONSAVEL_MAP:
        return RESPONSAVEL_MAP[nome]
    # normalized
    nome_n = _norm(nome)
    for k, v in RESPONSAVEL_MAP.items():
        if _norm(k) == nome_n:
            return v
    return None

def _para_date(valor):
    if valor is None or (isinstance(valor, float) and str(valor) == 'nan'):
        return None
    if isinstance(valor, datetime): return valor.date()
    if isinstance(valor, date):     return valor
    s = str(valor).strip().split()[0]
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d.%m.%Y'):
        try: return datetime.strptime(s, fmt).date()
        except ValueError: pass
    return None

_PAT_DATA = re.compile(r'(\d{1,2})[/\.](\d{1,2})[/\.](\d{2,4})')

def _data_fatal(descricao, ano_ref):
    res = []
    for m in _PAT_DATA.finditer(str(descricao or '')):
        d, mo, a_s = int(m.group(1)), int(m.group(2)), m.group(3)
        a = int(a_s)
        if a < 100:   a = (ano_ref // 100) * 100 + a
        elif len(a_s) == 3: a = int(str(ano_ref)[:1] + a_s)
        try: res.append(date(a, mo, d))
        except ValueError: pass
    return max(res) if res else None

def _d1_util(fatal, fer_np):
    prev = np.datetime64(fatal.strftime('%Y-%m-%d'), 'D') - np.timedelta64(1, 'D')
    return date.fromisoformat(str(
        np.busday_offset(prev, 0, roll='backward', holidays=fer_np)))

def _prox_util(ref, fer_np):
    return date.fromisoformat(str(
        np.busday_offset(np.datetime64(ref.strftime('%Y-%m-%d'), 'D'),
                         1, roll='forward', holidays=fer_np)))

TST_EXCECAO = 'protocolar recurso sobre decisao tst ou justificar nao recurso'

def _cor(descricao, justificativa, data_cad, hoje, fer_np):
    """
    Retorna (fill, fonte, cor_base, data_err, div_str).
    cor_base: categoria para contagem no RESUMO (verde/amarelo/rosa/vermelho).
    Laranjas são categorizadas pela data fatal, não pela cor.
    """
    desc = str(descricao or '')
    just = str(justificativa or '').strip()

    # 1. Verde forçado: audiências
    if 'AUDIENCIA DE JULGAMENTO' in _norm(desc) or 'ACOMPANHAR JULGAMENTO' in _norm(desc):
        return F_VERDE, FONTE_NORMAL, 'verde', False, ''

    # 2. Laranja
    tem_nao = bool(re.search(r'\bn[aã]o\b', desc, re.IGNORECASE))
    eh_tst  = TST_EXCECAO in _norm(desc)
    laranja  = bool(just) or (tem_nao and not eh_tst)

    # 3. Cor por data fatal
    ano_ref  = data_cad.year if data_cad else hoje.year
    fatal    = _data_fatal(desc, ano_ref)
    amanha   = hoje + timedelta(days=1)
    data_err, div_str = False, ''

    if fatal:
        d1_calc = _d1_util(fatal, fer_np)
        if data_cad and data_cad != d1_calc:
            data_err = True
            div_str  = (f'Cadastro {data_cad.strftime("%d/%m/%Y")} '
                        f'<> D-1 util {d1_calc.strftime("%d/%m/%Y")}')
        cor_base = ('vermelho' if fatal < hoje else
                    'rosa'     if fatal == hoje else
                    'amarelo'  if fatal == amanha else 'verde')
    elif data_cad:
        fatal_inf = _prox_util(data_cad, fer_np)
        cor_base  = ('vermelho' if fatal_inf < hoje else
                     'rosa'     if fatal_inf == hoje else
                     'amarelo'  if fatal_inf == amanha else 'verde')
    else:
        # sem data em nenhum campo: fatal = amanhã
        cor_base = 'amarelo'

    if laranja:
        return F_LARANJA, FONTE_NORMAL, cor_base, data_err, div_str

    FILLS  = {'verde': F_VERDE, 'amarelo': F_AMARELO,
              'rosa': F_ROSA,   'vermelho': F_VERMELHO}
    FONTES = {'vermelho': FONTE_BRANCA}
    return FILLS[cor_base], FONTES.get(cor_base, FONTE_NORMAL), cor_base, data_err, div_str


def _cnt(lista):
    c = dict(total=0, amarelo=0, rosa=0, vermelho=0, verde=0)
    for it in lista:
        c['total'] += 1
        c[it[5]] = c.get(it[5], 0) + 1   # it[5] = cor_base
    c['atraso'] = c['rosa'] + c['vermelho']
    return c

def _cel_r(ws, row, col, v=None, fill=None, font=None, align=None):
    """Célula do RESUMO — sem borda."""
    c = ws.cell(row, col, v)
    if fill: c.fill = fill
    if font: c.font = font
    c.alignment = align or AL_C
    return c

def _cel_d(ws, row, col, v=None, fill=None, font=None, align=None):
    """Célula de dados — com borda fina."""
    c = ws.cell(row, col, v)
    if fill: c.fill = fill
    else:    c.fill = F_NONE
    c.font   = font or FONTE_NORMAL
    c.alignment = align or AL_C
    c.border = BORDA_FINA
    return c

def _cabec_resumo(ws, row, headers):
    for j, h in enumerate(headers, 1):
        _cel_r(ws, row, j, h, F_CABEC, FONTE_CABEC, AL_C_WRAP)

# ─── FUNÇÃO PRINCIPAL ─────────────────────────────────────────────────────────
def gerar_relatorio(input_source):
    """
    input_source: bytes, str (caminho) ou file-like.
    Retorna: (output_bytes, resumo_dict)
    resumo_dict contém chaves extras para o painel de alertas:
      'alertas': lista de (tipo, descricao) para exibição no Streamlit
    """
    # Leitura
    if isinstance(input_source, bytes):
        buf = io.BytesIO(input_source)
    else:
        buf = input_source

    if _USE_PANDAS:
        df = pd.read_excel(buf, dtype=str)
        df.columns = [str(c).strip() for c in df.columns]
        # Normalizar nomes de colunas
        rename = {}
        for c in df.columns:
            cn = _norm(c)
            for alvo, variantes in [
                ('Data',            ['DATA']),
                ('Executante',      ['EXECUTANTE']),
                ('Natureza',        ['NATUREZA']),
                ('Tipo / Subtipo',  ['TIPO / SUBTIPO','TIPO/SUBTIPO','TIPO']),
                ('Descricao',       ['DESCRICAO','DESCRICAO']),
                ('Processo',        ['PROCESSO']),
                ('Cliente',         ['CLIENTE']),
                ('Contrario',       ['CONTRARIO']),
                ('Orgao / Cidade',  ['ORGAO / CIDADE','ORGAO/CIDADE','ORGAO']),
                ('Vara',            ['VARA']),
                ('Status',          ['STATUS']),
                ('Justificativa',   ['JUSTIFICATIVA']),
                ('Pasta',           ['PASTA']),
            ]:
                if cn in variantes or any(_norm(v) == cn for v in variantes):
                    rename[c] = alvo
        df = df.rename(columns=rename)
        header  = list(df.columns)
        rows_in = df.values.tolist()
    else:
        wb_in = openpyxl.load_workbook(buf, data_only=True)
        ws_in = wb_in.active
        all_rows = list(ws_in.iter_rows(values_only=True))
        header   = [str(h or '').strip() for h in all_rows[0]]
        rows_in  = [list(r) for r in all_rows[1:]]

    def _idx(nome):
        try: return header.index(nome)
        except ValueError:
            hn = _norm(nome)
            for i, h in enumerate(header):
                if _norm(h) == hn: return i
            return None

    idx_data = _idx('Data')
    idx_exec = _idx('Executante')
    idx_desc = _idx('Descricao')
    idx_just = _idx('Justificativa')
    idx_proc = _idx('Processo')

    # Mapas de coluna do input → coluna de saída
    mapa_entrada = {}
    for nome_s in NOMES_COLUNAS:
        i = _idx(nome_s)
        if i is not None:
            mapa_entrada[nome_s] = i

    hoje   = date.today()
    fer_np = _feriados_np([hoje.year, hoje.year + 1])

    # ── Processar linhas ──────────────────────────────────────────────────────
    processadas, nao_mapeados = [], []
    colunas_novas = []   # colunas do input que não estão no padrão

    for row in rows_in:
        if all(str(v or '').strip() in ('', 'nan') for v in row): continue

        def _v(i):
            if i is None or i >= len(row): return None
            v = row[i]
            return None if str(v or '').strip() in ('', 'nan', 'None') else v

        exec_raw = _v(idx_exec)
        desc     = str(_v(idx_desc) or '')
        just     = str(_v(idx_just) or '')
        data_cad = _para_date(_v(idx_data))

        supervisor = _mapear(exec_raw)
        fill, font, cor_base, data_err, div_str = _cor(
            desc, just, data_cad, hoje, fer_np)

        # Montar linha de saída com as 15 colunas padrão
        row_out = {}
        for nome_s, i_in in mapa_entrada.items():
            row_out[nome_s] = row[i_in] if i_in < len(row) else None
        # Campos calculados
        row_out['Responsavel']       = supervisor or ''
        row_out['Divergencia de Data'] = div_str or None
        # Formatar data
        if data_cad and row_out.get('Data'):
            raw_data = str(row_out['Data']).strip()
            # Manter hora se existir
            partes = str(_v(idx_data) or '').strip().split()
            hora = partes[1] if len(partes) > 1 else '00:00'
            row_out['Data'] = data_cad.strftime('%d/%m/%Y') + ' ' + hora[:5]

        # Checar colunas extras no input
        for i, h in enumerate(header):
            if h and _norm(h) not in [_norm(n) for n in NOMES_COLUNAS]:
                if h not in colunas_novas:
                    colunas_novas.append(h)

        if supervisor:
            processadas.append((row_out, supervisor, str(exec_raw or ''),
                                 fill, font, cor_base, data_err, div_str))
        else:
            proc_val = _v(idx_proc)
            nao_mapeados.append((str(exec_raw or ''), str(proc_val or '')))

    # Agrupamentos
    por_sup  = {}
    por_exec = {}
    for it in processadas:
        _, sup, exe, *_ = it
        por_sup.setdefault(sup, []).append(it)
        exe1 = exe.split('\n')[0].strip()
        por_exec.setdefault(exe1, []).append(it)

    # Laranjas para revisão
    laranjas   = [(it[0], it[1], it[2]) for it in processadas if it[3] == F_LARANJA]
    erros_data = [(it[0], it[1], it[2], it[7]) for it in processadas if it[6]]

    # ── Workbook ──────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ═══════════════════════════════════════════════════════════════════
    # ABA RESUMO
    # ═══════════════════════════════════════════════════════════════════
    ws_r = wb.create_sheet('RESUMO', 0)
    ws_r.column_dimensions['A'].width = 42
    ws_r.column_dimensions['B'].width = 38
    for col in 'CDEFGH':
        ws_r.column_dimensions[col].width = 16

    L = 1
    # Linha 1: título
    c_tit = ws_r.cell(L, 1, f'Relatorio de Prazos Preclusivos - {hoje.strftime("%d/%m/%Y")}')
    c_tit.font = FONTE_TITULO
    L += 2  # linha 3

    H_SUP = ['Responsavel','Total de Prazos','D-1','Fatal',
              'Prazo Vencido - Pendente de Baixa','Dentro do Prazo',
              'Total em Atraso (Fatal + Prazo Vencido - Pendente de Baixa)']
    H_EXE = ['Executante','Responsavel','Total de Prazos','D-1','Fatal',
              'Prazo Vencido - Pendente de Baixa','Dentro do Prazo',
              'Total em Atraso (Fatal + Prazo Vencido - Pendente de Baixa)']

    FILLS_COL_SUP  = [None, None, F_AMARELO, F_ROSA, F_VERMELHO, F_VERDE, None]
    FONTES_COL_SUP = [None, None, None,       None,   FONTE_BRANCA, None,  None]

    def _linha_sup(ws, L, label, c, is_total=False):
        vals = [label, c['total'],
                c.get('amarelo') or None, c.get('rosa') or None,
                c.get('vermelho') or None, c.get('verde') or None,
                c.get('atraso') or None]
        for j, (v, fc, ff) in enumerate(zip(vals, FILLS_COL_SUP, FONTES_COL_SUP), 1):
            cell = ws.cell(L, j, v)
            cell.alignment = AL_E if j == 1 else AL_C
            if is_total:
                cell.fill = F_TOTAL
                cell.font = FONTE_NEGRITO
                # manter cor das colunas coloridas mesmo no total
                if fc and v: cell.fill = fc
                if ff and v: cell.font = ff
            else:
                if fc and v: cell.fill = fc
                if ff and v: cell.font = ff

    # Tabela por supervisor
    _cabec_resumo(ws_r, L, H_SUP); L += 1
    tot = dict(total=0, amarelo=0, rosa=0, vermelho=0, verde=0, atraso=0)
    for sup in ORDEM_SUPERVISORES:
        if sup not in por_sup: continue
        c = _cnt(por_sup[sup])
        _linha_sup(ws_r, L, sup, c); L += 1
        for k in tot: tot[k] += c.get(k, 0)
    _linha_sup(ws_r, L, 'TOTAL GERAL', tot, is_total=True); L += 2

    # Tabela por executante
    ws_r.cell(L, 1, 'Prazos por Executante:').font = FONTE_NEGRITO; L += 1
    _cabec_resumo(ws_r, L, H_EXE); L += 1
    tot_e = dict(total=0, amarelo=0, rosa=0, vermelho=0, verde=0, atraso=0)

    FILLS_COL_EXE  = [None, None, None, F_AMARELO, F_ROSA, F_VERMELHO, F_VERDE, None]
    FONTES_COL_EXE = [None, None, None, None,       None,   FONTE_BRANCA, None,  None]

    for exe in sorted(por_exec.keys()):
        c = _cnt(por_exec[exe])
        sup = por_exec[exe][0][1]
        vals = [exe, sup, c['total'],
                c.get('amarelo') or None, c.get('rosa') or None,
                c.get('vermelho') or None, c.get('verde') or None,
                c.get('atraso') or None]
        for j, (v, fc, ff) in enumerate(zip(vals, FILLS_COL_EXE, FONTES_COL_EXE), 1):
            cell = ws_r.cell(L, j, v)
            if fc and v: cell.fill = fc
            if ff and v: cell.font = ff
            cell.alignment = AL_E if j <= 2 else AL_C
        L += 1
        for k in tot_e: tot_e[k] += c.get(k, 0)

    tot_e_vals = [None, 'TOTAL GERAL', tot_e['total'],
                  tot_e.get('amarelo') or None, tot_e.get('rosa') or None,
                  tot_e.get('vermelho') or None, tot_e.get('verde') or None,
                  tot_e.get('atraso') or None]
    for j, (v, fc, ff) in enumerate(zip(tot_e_vals, FILLS_COL_EXE, FONTES_COL_EXE), 1):
        cell = ws_r.cell(L, j, v)
        cell.fill = F_TOTAL; cell.font = FONTE_NEGRITO
        if fc and v: cell.fill = fc
        if ff and v: cell.font = ff
        cell.alignment = AL_E if j <= 2 else AL_C
    L += 2

    # ── Seção laranjas ────────────────────────────────────────────────
    if laranjas:
        ws_r.cell(L, 1,
            f'ATENCAO: {len(laranjas)} linha(s) em laranja para revisao:'
        ).font = Font(bold=True, size=11, color='FF6600'); L += 1

        H_LAR = ['Data','Responsavel','Executante','Processo','Cliente','Descricao','Motivo']
        _cabec_resumo(ws_r, L, H_LAR)
        for j in range(1, len(H_LAR) + 1):
            ws_r.cell(L, j).fill = F_LARANJA
            ws_r.cell(L, j).font = FONTE_NEGRITO
        L += 1

        for row_v, sup, exe in laranjas:
            just_v = str(row_v.get('Justificativa') or '').strip()
            desc_v = str(row_v.get('Descricao') or '')
            motivo = (f'Justificativa: {just_v[:80]}' if just_v
                      else 'Descricao contem NAO')
            vals = [
                row_v.get('Data', ''),
                sup,
                exe,
                row_v.get('Processo', ''),
                row_v.get('Cliente', ''),
                desc_v[:120],
                motivo,
            ]
            for j, v in enumerate(vals, 1):
                c2 = ws_r.cell(L, j, v)
                c2.fill = F_LARANJA
                c2.alignment = AL_E_WRAP
            L += 1
        # Larguras extras para as novas colunas do alerta
        ws_r.column_dimensions['E'].width = 35
        ws_r.column_dimensions['F'].width = 50
        ws_r.column_dimensions['G'].width = 40
        L += 1

    # ── Seção erros de data ───────────────────────────────────────────
    if erros_data:
        ws_r.cell(L, 1,
            f'ATENCAO: {len(erros_data)} data(s) fora do D-1 util:'
        ).font = Font(bold=True, size=11, color='CC0000'); L += 1

        H_ERR = ['Data Cadastrada','Responsavel','Executante','Processo',
                  'Cliente','Descricao','Divergencia (Cadastro x D-1 Util)']
        _cabec_resumo(ws_r, L, H_ERR)
        for j in range(1, len(H_ERR) + 1):
            ws_r.cell(L, j).fill = F_DATA_ERR
            ws_r.cell(L, j).font = FONTE_BRANCA
        L += 1

        for row_v, sup, exe, div in erros_data:
            desc_v = str(row_v.get('Descricao') or '')
            vals = [
                row_v.get('Data', ''),
                sup,
                exe,
                row_v.get('Processo', ''),
                row_v.get('Cliente', ''),
                desc_v[:120],
                div,
            ]
            for j, v in enumerate(vals, 1):
                c2 = ws_r.cell(L, j, v)
                c2.fill = PatternFill('solid', fgColor='FFE0E0')
                c2.alignment = AL_E_WRAP
            L += 1
        ws_r.column_dimensions['E'].width = max(35, ws_r.column_dimensions['E'].width)
        ws_r.column_dimensions['F'].width = max(50, ws_r.column_dimensions['F'].width)
        ws_r.column_dimensions['G'].width = max(40, ws_r.column_dimensions['G'].width)
        L += 1

    # ── Seção não mapeados ────────────────────────────────────────────
    if nao_mapeados:
        ws_r.cell(L, 1,
            f'ATENCAO: {len(nao_mapeados)} executante(s) nao mapeado(s) — '
            f'linhas excluidas do relatorio:'
        ).font = Font(bold=True, size=11, color='CC0000'); L += 1

        H_NM = ['Executante (nao reconhecido)','Numero do Processo',
                 'Acao necessaria']
        _cabec_resumo(ws_r, L, H_NM)
        for j in range(1, len(H_NM) + 1):
            ws_r.cell(L, j).fill = F_VERMELHO
            ws_r.cell(L, j).font = FONTE_BRANCA
        L += 1

        for exe, proc in nao_mapeados:
            vals = [exe, proc,
                    'Comunicar gestao do sistema para inclusao no cadastro']
            for j, v in enumerate(vals, 1):
                c2 = ws_r.cell(L, j, v)
                c2.fill = PatternFill('solid', fgColor='FFE0E0')
                c2.alignment = AL_E_WRAP
            L += 1
        L += 1

    # Legenda
    ws_r.cell(L, 1, 'LEGENDA:').font = FONTE_NEGRITO; L += 1
    for fill, font, txt in [
        (F_VERDE,    FONTE_NORMAL,  'Dentro do prazo (D-1 ainda nao chegou)'),
        (F_AMARELO,  FONTE_NORMAL,  'D-1 (hoje e o dia de cumprir, fatal amanha)'),
        (F_ROSA,     FONTE_NORMAL,  'Fatal hoje (atrasado - prazo vence hoje)'),
        (F_VERMELHO, FONTE_BRANCA,  'Fatal ja passou (perda de prazo)'),
        (F_LARANJA,  FONTE_NORMAL,  'Com justificativa ou NAO na descricao'),
        (F_DATA_ERR, FONTE_BRANCA,  'Celula Data: cadastro fora do D-1 util'),
    ]:
        c2 = ws_r.cell(L, 1, txt)
        c2.fill = fill; c2.font = font; c2.alignment = AL_E; L += 1

    # ═══════════════════════════════════════════════════════════════════
    # ABAS POR SUPERVISOR
    # ═══════════════════════════════════════════════════════════════════
    for sup in ORDEM_SUPERVISORES:
        if sup not in por_sup: continue
        ws = wb.create_sheet(sup[:31])

        # Larguras
        for j, (nome, larg) in enumerate(COLUNAS_SAIDA, 1):
            ws.column_dimensions[get_column_letter(j)].width = larg

        # Freeze pane e autofiltro
        ws.freeze_panes = 'A2'

        # Cabeçalho linha 1
        for j, (nome, _) in enumerate(COLUNAS_SAIDA, 1):
            _cel_d(ws, 1, j, nome, F_CABEC, FONTE_CABEC, AL_C_WRAP)

        # Ordenar por Data
        itens = por_sup[sup]
        def _sort_key(it):
            data_str = it[0].get('Data', '') or ''
            try: return datetime.strptime(str(data_str).strip()[:10], '%d/%m/%Y')
            except: return datetime.min
        itens_sorted = sorted(itens, key=_sort_key)

        # Dados
        for li, it in enumerate(itens_sorted, 2):
            row_v, _, _, fill, font, _, data_err, _ = it
            for j, nome_col in enumerate(NOMES_COLUNAS, 1):
                v = row_v.get(nome_col)
                disp = None if str(v or '').strip() in ('nan','None','') else v
                al = AL_E_WRAP if j == 1 else AL_C_WRAP
                c2 = _cel_d(ws, li, j, disp, fill, font, al)
                # Flag data divergente: sobrepõe cor da linha
                if nome_col == 'Data' and data_err:
                    c2.fill = F_DATA_ERR
                    c2.font = FONTE_BRANCA

        # Autofiltro (após preencher dados)
        ultima_col = get_column_letter(len(COLUNAS_SAIDA))
        ws.auto_filter.ref = f'A1:{ultima_col}1'

    # ── Retorno ───────────────────────────────────────────────────────
    # Montar lista de alertas para o painel Streamlit
    alertas = []
    if nao_mapeados:
        alertas.append(('NAO_MAPEADO', nao_mapeados))
    if erros_data:
        alertas.append(('DATA_DIVERGENTE', [
            (r[1], r[2], r[3]) for r in erros_data  # sup, exe, div
        ]))
    if colunas_novas:
        alertas.append(('COLUNA_NOVA', colunas_novas))

    cnt_total = _cnt(processadas)
    resumo = {
        'total':          cnt_total['total'],
        'nao_mapeados':   len(nao_mapeados),
        'data_erros':     len(erros_data),
        'laranjas':       len(laranjas),
        'por_supervisor': {s: _cnt(v) for s, v in por_sup.items()},
        'alertas':        alertas,
        'nao_mapeados_lista': nao_mapeados,
        'erros_data_lista':   [(r[1], r[2], r[3]) for r in erros_data],
        'colunas_novas':      colunas_novas,
    }

    out = io.BytesIO()
    wb.save(out); out.seek(0)
    return out.getvalue(), resumo


# ─── EXECUÇÃO DIRETA ─────────────────────────────────────────────────────────
if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('Uso: python gerar_relatorio_prazos.py <entrada.xlsx> [saida.xlsx]')
        sys.exit(1)
    entrada = sys.argv[1]
    if len(sys.argv) > 2:
        saida = sys.argv[2]
    else:
        saida = f'Relatorio_Prazos_{date.today().strftime("%d-%m-%Y")}.xlsx'
    out_bytes, res = gerar_relatorio(entrada)
    with open(saida, 'wb') as f: f.write(out_bytes)
    print(f'\nSalvo: {saida}')
    print(f"Total: {res['total']} | Não mapeados: {res['nao_mapeados']} | "
          f"Erros data: {res['data_erros']} | Laranjas: {res['laranjas']}")
    for tipo, dados in res['alertas']:
        if tipo == 'NAO_MAPEADO':
            print(f'\n⚠️  EXECUTANTES NÃO MAPEADOS ({len(dados)}):')
            for exe, proc in dados: print(f'   - {exe} | Processo: {proc}')
        elif tipo == 'DATA_DIVERGENTE':
            print(f'\n⚠️  DATAS FORA DO D-1 ({len(dados)}):')
            for sup, exe, div in dados: print(f'   - {exe} ({sup}): {div}')
        elif tipo == 'COLUNA_NOVA':
            print(f'\n⚠️  COLUNAS NÃO PREVISTAS NA DOCUMENTAÇÃO: {dados}')
